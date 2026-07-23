#!/usr/bin/env python3
"""
MHLW ICD-10 2013 版 (基本分類表) を FHIR CodeSystem (content=complete) に変換。

Source:
  - 厚労省 統計情報 https://www.mhlw.go.jp/toukei/sippei/
  - ファイル: kihon2013.xlsx (基本分類表、15,586 concept)
  - License: 公共データ利用規約 (PDL 1.0)、出典明記条件下で再配布可

Output:
  - `tx-server-build/mhlw-icd10-src/CodeSystem-mhlw-icd10-2013-full.json`
    - FHIR R4 CodeSystem, `content: complete`
    - jpfhir-terminology の fragment CS と同 URL (`http://jpfhir.jp/fhir/core/mhlw/CodeSystem/ICD10-2013-full`)
    - fhirserver 側で fragment を上書き

Notes:
  - Code format は「表示用コード」列の値そのまま (例: A00, A000, A00-A09, I)
  - dot なしの concatenated 形式 (A00.0 → A000) が MHLW form
  - Chapter は Latin uppercase Roman (I, II, ..., XXII) を採番

Usage:
  ./scripts/build-mhlw-icd10-full.py
"""
from __future__ import annotations
import json, sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = REPO_ROOT / "tx-server-build/mhlw-icd10-src/kihon2013.xlsx"
OUT_JSON = REPO_ROOT / "tx-server-build/mhlw-icd10-src/CodeSystem-mhlw-icd10-2013-full.json"

CS_URL = "http://jpfhir.jp/fhir/core/mhlw/CodeSystem/ICD10-2013-full"
CS_VERSION = "1.1.1"  # jpfhir-terminology 2.2606.0 の version と合わせる
CS_ID = "mhlw-codesystem-icd10-2013-jp-full"

# Chapter roman numeral mapping (章番号 "01"..."22" → Roman)
ROMAN = {
    1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII",
    9: "IX", 10: "X", 11: "XI", 12: "XII", 13: "XIII", 14: "XIV", 15: "XV",
    16: "XVI", 17: "XVII", 18: "XVIII", 19: "XIX", 20: "XX", 21: "XXI", 22: "XXII",
}


def main() -> None:
    if not SRC_XLSX.exists():
        print(f"ERROR: source file not found: {SRC_XLSX}", file=sys.stderr)
        print("Download: curl -o tx-server-build/mhlw-icd10-src/kihon2013.xlsx \\", file=sys.stderr)
        print("            https://www.mhlw.go.jp/toukei/sippei/xls/kihon2013.xlsx", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {SRC_XLSX}...", file=sys.stderr)
    wb = load_workbook(SRC_XLSX, read_only=True, data_only=True)
    ws = wb["基本分類表データ"]

    concepts: list[dict] = []
    seen_codes: set[str] = set()
    dup_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # columns: 種別, 分類単位, 章番号, 中間分類, 三桁分類, コード, 剣星, 表示用コード, 注・細分類等, コード名, 大分類, 中分類, 小分類, 死因分類
        (kind, _, chap_no, mid, three, code, _, disp_code, _, name, _, _, _, _) = row
        if not (kind and disp_code):
            continue

        code_str = str(disp_code).strip()
        name_str = (name or "").strip()
        if not code_str or not name_str:
            continue

        # Chapter: xlsx 列は "第Ⅰ章"、表示用コードは Roman numeral の Latin 表記に置換
        if kind == "01.章":
            try:
                ch_num = int(str(chap_no))
                code_str = ROMAN.get(ch_num, code_str)
            except (TypeError, ValueError):
                pass

        if code_str in seen_codes:
            dup_count += 1
            continue
        seen_codes.add(code_str)

        concept: dict = {"code": code_str, "display": name_str}
        concepts.append(concept)

    if dup_count > 0:
        print(f"WARNING: {dup_count} duplicate 表示用コード skipped", file=sys.stderr)

    print(f"Extracted {len(concepts)} unique concepts", file=sys.stderr)

    cs: dict = {
        "resourceType": "CodeSystem",
        "id": CS_ID,
        "url": CS_URL,
        "version": CS_VERSION,
        "name": "MHLW_codeSystem_icd10_2013_jp_full",
        "title": "MHLW 厚労省 ICD-10 2013 版 (完全版)",
        "status": "active",
        "experimental": False,
        "date": date.today().isoformat(),
        "publisher": "厚生労働省 (data provider) / fhir-jp-validator (converter)",
        "description": (
            "厚労省 疾病、傷害及び死因の統計分類 (2013 年基本分類表) の全 15,000+ concept を "
            "FHIR CodeSystem (content=complete) 形式に変換したもの。\n\n"
            "Source: https://www.mhlw.go.jp/toukei/sippei/ の kihon2013.xlsx\n"
            "License: 公共データ利用規約 (PDL 1.0)、出典明記条件下で再配布可\n"
            "出典: 『疾病、傷害及び死因の統計分類』(厚生労働省) を加工して作成\n\n"
            "このリソースは jpfhir-terminology 2.2606.0 の fragment CS "
            "(先頭 2,000 concept のみ) を complete で上書きする目的で fhir-jp-validator "
            "の scripts/build-mhlw-icd10-full.py で生成。"
        ),
        "copyright": (
            "『疾病、傷害及び死因の統計分類』(厚生労働省) を加工して作成\n"
            "本 CodeSystem 自体の再配布は 公共データ利用規約 (PDL 1.0) に準拠。"
        ),
        "caseSensitive": True,
        "content": "complete",
        "count": len(concepts),
        "language": "ja",
        "concept": concepts,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(cs, f, ensure_ascii=False, indent=2)

    size_mb = OUT_JSON.stat().st_size / 1024 / 1024
    print(f"Written {OUT_JSON} ({size_mb:.1f} MB, {len(concepts)} concepts)", file=sys.stderr)


if __name__ == "__main__":
    main()
